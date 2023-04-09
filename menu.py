import openpyxl

#read excel using 
dataframe = openpyxl.load_workbook('itemlist.xlsx')
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
# 1st argument is the row number
def Get_Menu():
    item_id=[]
    item_name=[]
    item_price=[]
    item_description=[]

    #Get the item id in column 1
    for col in dataframe1.iter_cols(1, 1):
    
        for row in range(1, dataframe1.max_row): 
            a=col[row].value
            item_id.append(a)

    #get the item name in column 2
    for col in dataframe1.iter_cols(2, 2):

    
        for row in range(1, dataframe1.max_row): 
            a=col[row].value
            item_name.append(a)
    
    #get the item price in column 3
    for col in dataframe1.iter_cols(3, 3):

    
        for row in range(1, dataframe1.max_row): 
            a=col[row].value
            item_price.append(a)
    
    #Get the item description column 4
    for col in dataframe1.iter_cols(4, 4):
        #print(row)
    
        for row in range(1, dataframe1.max_row): 
            #print(col[row].value)
            a=col[row].value
            item_description.append(a)
    return item_id, item_name, item_price, item_description

def Display_Menu(item_id, item_name, item_price, item_description):
    print("---------------------------------------------------------------------")
    print("\t\t\t\tMENU")
    print("---------------------------------------------------------------------")
    print("ID \tName\t\t\t\t\t\t\t\tPrice\t\t")
    for i in range(len(item_id)):
        print(item_id[i], end='\t')
        print(item_name[i].ljust(63), end='')
        print("P"+str(item_price[i])) 
        
